# --------------------------------------------------------------

import fitz  # PyMuPDF
import pytesseract
from PIL import Image
from io import BytesIO
import pdfplumber
import os
from docx import Document
import openpyxl
from typing import List, Dict, Any, Optional
import pytesseract

#pytesseract.pytesseract.tesseract_cmd = '/opt/homebrew/bin/tesseract'


# --------------------------------------------------------------
class DocumentExtractor:
    """A class to handle document text extraction for various file formats."""
    
    def __init__(self, tesseract_path: Optional[str] = None):
        """
        Initialize the DocumentExtractor.
        
        Args:
            tesseract_path (str, optional): Path to tesseract executable
        """
        if tesseract_path:
            pytesseract.pytesseract.tesseract_cmd = tesseract_path
    
    def _extract_images_from_pdf_page(self, pdf_file: str, page_num: int) -> List[Image.Image]:
        """
        Extract images from a single PDF page using PyMuPDF.
        
        Args:
            pdf_file (str): Path to the PDF file
            page_num (int): Page number to extract images from
            
        Returns:
            List[Image.Image]: List of extracted images
        """
        images = []
        doc = fitz.open(pdf_file)
        page = doc.load_page(page_num)
        image_list = page.get_images(full=True)
        
        for img_index, img in enumerate(image_list):
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            image = Image.open(BytesIO(image_bytes))
            images.append(image)
        
        doc.close()
        return images

    def _ocr_on_image(self, image: Image.Image) -> str:
        """
        Perform OCR on the provided image using pytesseract.
        
        Args:
            image (Image.Image): Image to perform OCR on
            
        Returns:
            str: Extracted text from the image
        """
        return pytesseract.image_to_string(image)

    def extract_text_from_pdf(self, pdf_file: str) -> str:
        """
        Extract structured text from a PDF file using pdfplumber.
        
        Args:
            pdf_file (str): Path to the PDF file
            
        Returns:
            str: Extracted text including regular text, tables, and OCR results
        """
        text_parts = []
        
        with pdfplumber.open(pdf_file) as pdf:
            for page_num in range(len(pdf.pages)):
                page = pdf.pages[page_num]
                
                # Extract regular text
                text_parts.append(page.extract_text() or '')
                
                # Extract and format tables
                tables = page.extract_tables()
                for table in tables:
                    text_parts.append("\nTable Data:")
                    for row in table:
                        text_parts.append("\t".join(str(cell) for cell in row if cell))
                
                # Extract and process images
                images = self._extract_images_from_pdf_page(pdf_file, page_num)
                for img_index, img in enumerate(images):
                    print(f"Performing OCR on Image {img_index + 1} on Page {page_num + 1}...")
                    image_text = self._ocr_on_image(img)
                    if image_text.strip():
                        text_parts.append(f"\nText from Image {img_index + 1} on Page {page_num + 1}:")
                        text_parts.append(image_text)
        
        return "\n".join(text_parts)

    def extract_text_from_docx(self, docx_file: str) -> str:
        """
        Extract text from a Word (DOCX) file using python-docx.
        
        Args:
            docx_file (str): Path to the DOCX file
            
        Returns:
            str: Extracted text from the document
        """
        doc = Document(docx_file)
        return "\n".join(para.text for para in doc.paragraphs)

    def extract_text_from_xlsx(self, xlsx_file: str) -> str:
        """
        Extract text from an Excel (XLSX) file using openpyxl.
        
        Args:
            xlsx_file (str): Path to the XLSX file
            
        Returns:
            str: Extracted text from the spreadsheet
        """
        wb = openpyxl.load_workbook(xlsx_file)
        text_parts = []
        
        for sheet in wb.sheetnames:
            active_sheet = wb[sheet]
            text_parts.append(f"\nSheet: {sheet}")
            for row in active_sheet.iter_rows(values_only=True):
                text_parts.append("\t".join(str(cell) for cell in row if cell))
        
        return "\n".join(text_parts)

    def extract_text(self, file_path: str) -> str:
        """
        Extract text from a document based on its file type.
        
        Args:
            file_path (str): Path to the document
            
        Returns:
            str: Extracted text from the document
        """
        file_extension = os.path.splitext(file_path)[1].lower()
        
        extractors = {
            '.pdf': self.extract_text_from_pdf,
            '.docx': self.extract_text_from_docx,
            '.xlsx': self.extract_text_from_xlsx
        }
        
        extractor = extractors.get(file_extension)
        if not extractor:
            raise ValueError(f"Unsupported file type: {file_extension}")
            
        print(f"Processing {file_extension} document: {file_path}")
        return extractor(file_path)

def load_document(file_path: str) -> str:
    """
    Load and extract text from a document.
    
    Args:
        file_path (str): Path to the document
        
    Returns:
        str: Extracted text from the document
    """
    tesseract_path = '/opt/homebrew/bin/tesseract'
    extractor = DocumentExtractor(tesseract_path=tesseract_path)
    try:
        extracted_text = extractor.extract_text(file_path)
        return extracted_text
    except Exception as e:
        print(f"Error extracting text: {str(e)}")
        return ""